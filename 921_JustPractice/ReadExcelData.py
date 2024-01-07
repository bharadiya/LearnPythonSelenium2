import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\poona\\PycharmProjects\\LearnSelenium\\ExcelFiles\\Test.xlsx")
sheets = wb.sheetnames
# print(wb.active.title)

sh = wb["Sheet1"]
print(type(sh))

# number 1 way of reading excel data
data = sh['B21'].value
print(data)

# number 2 way of reading excel data
data2 = sh.cell(1, 1).value
print(data2)

row = sh.max_row
column = sh.max_column
print(sh.max_row)
# 23
print(sh.max_column)
# 2

for i in range(1, row + 1):
    for j in range(1, column + 1):
        data = [(sh.cell(i, j)).value]
        print(data)
