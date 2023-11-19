import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\poona\\PycharmProjects\\LearnSelenium\\ExcelFiles\\Test.xlsx")
sheets = wb.sheetnames
print(wb.active.title)

sh1 = wb['Sheet1']
print(type(sh1))

data = sh1['A2'].value
print(data)

data2 = sh1.cell(2, 2).value
print(data2)

row = sh1.max_row
column = sh1.max_column
print(sh1.max_row)
print(sh1.max_column)

for i in range(1, row + 1):
    for j in range(1, column + 1):
        data = [(sh1.cell(i,j)).value]
        print(data)
