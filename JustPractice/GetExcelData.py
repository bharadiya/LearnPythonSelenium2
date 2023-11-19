import openpyxl

path = "C:\\Users\\poona\\PycharmProjects\\LearnSelenium\\ExcelFiles\\Test.xlsx"
wb = openpyxl.load_workbook(path)
# print(type(wb))

sheets = wb.sheetnames
print(type(sheets))
print(sheets)
print("Currently active sheet is " + wb.active.title)

sh = wb['Sheet1']
print(type(sh))

data = sh['A13'].value
print(data)

data2 = sh.cell(2, 1).value
print(data2)

row = sh.max_row
column = sh.max_column

print(row)
print(column)

for i in range(1, row + 1):
    for j in range(1, column + 1):
        print(sh.cell(i, j).value)
    print()
